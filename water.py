"""Domestic water consumption + hot-water (DHW) energy model.

Reads the monthly water consumption from the "Wasser" sheet of
source_data/Strom_Wasser_Verbrauch.xlsx (column C = monthly m³, from row 31)
and builds a typical monthly profile by averaging each calendar month over the
available years. From that we estimate the hot-water share and the energy
needed to heat it.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl

DEFAULT_WATER_XLSX = (Path(__file__).with_name("source_data")
                      / "Strom_Wasser_Verbrauch.xlsx")
DEFAULT_WATER_JSON = (Path(__file__).with_name("source_data")
                      / "water_monthly_default.json")

# Specific heat of water ≈ 1.163 kWh per m³ per K (4.186 kJ/(kg·K)).
WATER_KWH_PER_M3_K = 1.163


def _load_monthly_water_json(path: Path) -> dict[int, float]:
    import json

    data = json.loads(path.read_text(encoding="utf-8"))
    return {int(k): float(v) for k, v in data.items()}


def load_monthly_water_m3(path: Path = DEFAULT_WATER_XLSX, sheet: str = "Wasser",
                          start_row: int = 31, date_col: int = 1,
                          consumption_col: int = 3) -> dict[int, float]:
    """Typical monthly water use [m³], averaged per calendar month over years."""
    if not path.exists():
        if DEFAULT_WATER_JSON.exists():
            return _load_monthly_water_json(DEFAULT_WATER_JSON)
        raise FileNotFoundError(
            f"Water profile not found: {path} (and no {DEFAULT_WATER_JSON.name})"
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    buckets: dict[int, list] = defaultdict(list)
    for r in range(start_row, ws.max_row + 1):
        date = ws.cell(r, date_col).value
        cons = ws.cell(r, consumption_col).value
        if date is None or cons is None:
            continue
        buckets[date.month].append(float(cons))
    return {m: sum(v) / len(v) for m, v in sorted(buckets.items())}


def hot_water_energy_kwh(monthly_water_m3: dict[int, float], cfg: dict) -> dict:
    """Monthly hot-water *thermal* energy [kWh] from the water profile.

    A seasonal hot-water fraction is applied: in summer the total water use
    includes garden/outdoor water that is not heated, so the hot share is lower.
    """
    inlet = float(cfg.get("cold_inlet_c", 10.0))
    target = float(cfg.get("target_temp_c", 50.0))
    frac_winter = float(cfg.get("hot_fraction_winter", 0.40))
    frac_summer = float(cfg.get("hot_fraction_summer", 0.30))
    summer = set(cfg.get("summer_months", [5, 6, 7, 8, 9]))

    out = {}
    for m, total_m3 in monthly_water_m3.items():
        frac = frac_summer if m in summer else frac_winter
        hot_m3 = total_m3 * frac
        heat_kwh = hot_m3 * WATER_KWH_PER_M3_K * (target - inlet)
        out[m] = {"total_m3": total_m3, "hot_m3": hot_m3,
                  "fraction": frac, "heat_kwh": heat_kwh}
    return out


if __name__ == "__main__":
    prof = load_monthly_water_m3()
    print("Typical monthly water use (m³), averaged over years:")
    for m, v in prof.items():
        print(f"  month {m:>2}: {v:6.2f} m³")
    print(f"  annual total: {sum(prof.values()):.1f} m³")
